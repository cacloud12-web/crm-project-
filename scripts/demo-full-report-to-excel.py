#!/usr/bin/env python3
"""Convert demo-full-report.json to a multi-sheet Excel workbook."""

import json
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip3 install openpyxl")
    sys.exit(1)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "", name)[:28].strip() or "Employee"
    candidate = base
    i = 1
    while candidate in used:
        suffix = f"_{i}"
        candidate = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(candidate)
    return candidate


def style_header_row(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def write_table(ws, start_row, headers, rows, widths=None):
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header_row(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.alignment = CENTER if c > 1 else LEFT
        r += 1
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return r


def main():
    root = Path(__file__).resolve().parent.parent
    json_path = Path(sys.argv[1]) if len(sys.argv) > 1 else root / "storage/app/audits/demo-full-report.json"
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else root / "storage/app/audits/Demo_Full_Employee_Report.xlsx"

    if not json_path.exists():
        print(f"Missing JSON: {json_path}")
        print("Run on production first:")
        print("  /opt/alt/php83/usr/bin/php scripts/demo-full-report-export.php 20")
        sys.exit(1)

    data = json.loads(json_path.read_text(encoding="utf-8"))
    wb = Workbook()
    used_names = set()

    # --- All Employees Summary ---
    ws = wb.active
    ws.title = "All Employees"
    ws["A1"] = "Demo Report — All Employees"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws.merge_cells("A1:L1")
    ws["A2"] = f"Period: {data['from']} to {data['to']}  |  Generated: {data.get('generated_at', '')}"
    ws.merge_cells("A2:L2")

    headers = [
        "Employee", "Role", "Status", "Total Demos", "Still Open", "Completed",
        "Rescheduled", "Cancelled", "Missed", "Not Interested", "Outcomes Recorded",
        "Interested", "Thinking", "Purchased", "Purchasing", "Hold/Next",
    ]
    rows = []
    for emp in data["employees"]:
        s = emp["summary"]
        ob = s.get("outcome_breakdown", {})
        hold_next = (
            ob.get("Hold", 0) + ob.get("Next Week", 0) + ob.get("Next Month", 0)
            + ob.get("Left in between", 0)
        )
        rows.append([
            emp["employee_name"], emp.get("role", ""), emp.get("employee_status", ""),
            s["total_demos"], s["still_open"], s["completed"], s["rescheduled"],
            s["cancelled"], s["missed"], s["not_interested"], s["outcomes_recorded"],
            ob.get("Interested", 0), ob.get("Thinking", 0), ob.get("Purchased", 0),
            ob.get("Purchasing", 0), hold_next,
        ])

    gt = data.get("grand_totals", {})
    rows.append([
        "GRAND TOTAL", "", "", gt.get("total_demos", 0), gt.get("still_open", 0),
        gt.get("completed", 0), gt.get("rescheduled", 0), gt.get("cancelled", 0),
        gt.get("missed", 0), gt.get("not_interested", 0), "", "", "", "", "", "",
    ])

    end = write_table(ws, 4, headers, rows, [22, 12, 10, 12, 12, 12, 12, 12, 10, 14, 16, 12, 12, 12, 12, 12])
    for c in range(1, len(headers) + 1):
        ws.cell(row=end, column=c).font = BOLD
        ws.cell(row=end, column=c).fill = SUB_FILL
    ws.freeze_panes = "A5"

    detail_headers = [
        "Demo ID", "Firm Name", "CA Name", "Mobile", "Demo Date/Time", "Scheduled On",
        "Status", "Outcome", "Outcome Recorded", "Demo Provider", "Team Size", "Lead Status", "Notes",
    ]

    for emp in data["employees"]:
        s = emp["summary"]
        if s["total_demos"] == 0 and s["outcomes_recorded"] == 0:
            continue

        name = sheet_name(emp["employee_name"], used_names)
        ws_e = wb.create_sheet(name)

        ws_e["A1"] = f"Employee: {emp['employee_name']}"
        ws_e["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws_e.merge_cells("A1:M1")
        ws_e["A2"] = f"Email: {emp.get('email', '')}  |  Role: {emp.get('role', '')}  |  Status: {emp.get('employee_status', '')}"
        ws_e.merge_cells("A2:M2")

        summary_labels = [
            ("Total Demos", s["total_demos"]),
            ("Still Open", s["still_open"]),
            ("Completed", s["completed"]),
            ("Rescheduled", s["rescheduled"]),
            ("Cancelled", s["cancelled"]),
            ("Missed", s["missed"]),
            ("Not Interested", s["not_interested"]),
            ("Outcomes Recorded", s["outcomes_recorded"]),
        ]
        r = 4
        for label, val in summary_labels:
            ws_e.cell(row=r, column=1, value=label).font = BOLD
            ws_e.cell(row=r, column=2, value=val)
            r += 1

        ob = s.get("outcome_breakdown", {})
        if any(ob.values()):
            r += 1
            ws_e.cell(row=r, column=1, value="Outcome Breakdown").font = BOLD
            r += 1
            for outcome, count in ob.items():
                if count:
                    ws_e.cell(row=r, column=1, value=outcome)
                    ws_e.cell(row=r, column=2, value=count)
                    r += 1

        r += 1
        ws_e.cell(row=r, column=1, value="Demo Line Items").font = Font(bold=True, size=12)
        r += 1

        detail_rows = []
        for d in emp.get("demos", []):
            detail_rows.append([
                d.get("demo_id"), d.get("firm_name"), d.get("ca_name"), d.get("mobile_no"),
                d.get("demo_at"), d.get("scheduled_on"), d.get("status"), d.get("outcome"),
                d.get("outcome_recorded_at"), d.get("demo_provider"), d.get("team_size"),
                d.get("lead_status"), d.get("notes") or d.get("outcome_notes"),
            ])

        if not detail_rows:
            ws_e.cell(row=r, column=1, value="No demo line items in selected period.")
        else:
            write_table(ws_e, r, detail_headers, detail_rows, [10, 28, 22, 16, 18, 18, 12, 16, 18, 18, 10, 14, 30])
        ws_e.freeze_panes = "A5"

    wb.save(out_path)
    print(str(out_path))


if __name__ == "__main__":
    main()
