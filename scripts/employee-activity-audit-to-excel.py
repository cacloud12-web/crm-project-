#!/usr/bin/env python3
"""Convert employee-activity-audit.json to Excel."""

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip3 install openpyxl")
    sys.exit(1)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "", name)[:28].strip() or "Employee"
    candidate = base
    i = 1
    while candidate in used:
        suffix = f"_{i}"
        candidate = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(candidate)
    return candidate


def style_header_row(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def write_table(ws, start_row, headers, rows, widths=None):
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header_row(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.alignment = CENTER if c > 1 else LEFT
            if isinstance(val, str) and val.startswith("FLAG:"):
                cell.fill = WARN_FILL
        r += 1
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return r


def main():
    root = Path(__file__).resolve().parent.parent
    json_path = Path(sys.argv[1]) if len(sys.argv) > 1 else root / "storage/app/audits/employee-activity-audit.json"
    data = json.loads(json_path.read_text(encoding="utf-8")) if json_path.exists() else None
    if data is None:
        print(f"Missing JSON: {json_path}")
        sys.exit(1)

    period = data.get("period")
    if isinstance(period, dict):
        period_slug = f"{period.get('from')}_to_{period.get('to')}"
        default_name = f"CRM_Full_Report_Last_{period.get('days', '')}_Days_{period_slug}.xlsx"
    else:
        default_name = "CRM_Full_Report_All_Time.xlsx"
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else root / "storage/app/audits" / default_name
    wb = Workbook()
    used = set()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    period = data.get("period")
    period_txt = "ALL TIME" if period == "all_time" else f"{period.get('from')} to {period.get('to')}"
    ws["A1"] = "CRM Full Employee Report"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws.merge_cells("A1:R1")
    ws["A2"] = f"Period: {period_txt} | Generated: {data.get('generated_at', '')}"
    ws.merge_cells("A2:R2")
    ws["A3"] = "Demo achieved = demos SCHEDULED per day (daily target rule). Integrity issues: {}".format(data.get('integrity_issue_count', 0))
    ws.merge_cells("A3:R3")

    headers = [
        "Employee", "Leads Assigned (now)", "Leads Assigned (period)", "Demo Target (period)",
        "Demo Achieved (period)", "Achievement %", "Today Target", "Today Achieved",
        "Demos Scheduled", "Demos Completed", "Demos Open", "Follow-ups Total",
        "Demo Scheduled FU (open)", "Demo Completed FU", "Calls Total", "Purchases",
        "Purchased Demo Results", "Integrity Issues",
    ]
    rows = []
    for emp in data["employees"]:
        s = emp["summary"]
        rows.append([
            emp["employee_name"],
            s.get("leads_assigned_active", 0),
            s.get("leads_assigned_in_period", 0),
            s.get("demo_target_period", 0),
            s.get("demo_achieved_period", 0),
            s.get("demo_achievement_pct", 0),
            s.get("demo_target_today", 0),
            s.get("demo_achieved_today", 0),
            s["demos_scheduled_created"],
            s["demos_completed"],
            s["demos_still_open"],
            s["followups_total"],
            s["followups_open_demo_scheduled"],
            s["followups_demo_completed"],
            s["calls_total"],
            s["purchases_total"],
            s["purchased_demo_results"],
            s.get("integrity_issue_count", 0),
        ])
    gt = data.get("grand_totals", {})
    rows.append([
        "GRAND TOTAL",
        gt.get("leads_assigned_active", 0),
        gt.get("leads_assigned_in_period", 0),
        gt.get("demo_target_period", 0),
        gt.get("demo_achieved_period", 0),
        "",
        "",
        "",
        gt.get("demos_scheduled_created", 0),
        gt.get("demos_completed", 0),
        "",
        gt.get("followups_total", 0),
        "",
        "",
        gt.get("calls_total", 0),
        gt.get("purchases_total", 0),
        gt.get("purchased_demo_results", 0),
        data.get("integrity_issue_count", 0),
    ])
    end = write_table(ws, 5, headers, rows, [22, 16, 16, 14, 16, 12, 12, 14, 14, 14, 12, 14, 18, 16, 12, 12, 18, 14])
    for c in range(1, len(headers) + 1):
        ws.cell(row=end, column=c).font = BOLD

    # Integrity
    if data.get("integrity_issues"):
        wi = wb.create_sheet("Integrity Issues")
        ih = ["Employee", "Type", "Record ID", "CA ID", "Firm", "Issues"]
        ir = []
        for issue in data["integrity_issues"]:
            ir.append([
                issue.get("employee"),
                issue.get("type"),
                issue.get("id"),
                issue.get("ca_id"),
                issue.get("firm_name") or issue.get("followup_type", ""),
                ", ".join(issue.get("issues", [])),
            ])
        write_table(wi, 1, ih, ir, [20, 14, 12, 10, 28, 40])

    detail_sets = [
        ("demos", ["demo_id", "firm_name", "ca_name", "mobile_no", "demo_at", "scheduled_on", "status", "outcome", "integrity_flags"]),
        ("followups", ["followup_id", "firm_name", "followup_type", "status", "scheduled_date", "remarks", "integrity_flags"]),
        ("calls", ["call_id", "firm_name", "called_at", "call_status", "call_note", "integrity_flags"]),
        ("purchases", ["purchase_id", "firm_name", "purchase_date", "software_name", "status", "integrity_flags"]),
    ]

    for emp in data["employees"]:
        name = sheet_name(emp["employee_name"], used)
        ws_e = wb.create_sheet(name)
        r = 1
        ws_e.cell(row=r, column=1, value=emp["employee_name"]).font = BOLD
        r += 2
        for section, cols in detail_sets:
            items = emp.get(section, [])
            ws_e.cell(row=r, column=1, value=section.title()).font = BOLD
            r += 1
            if not items:
                ws_e.cell(row=r, column=1, value="(none)")
                r += 2
                continue
            headers = [c.replace("_", " ").title() for c in cols]
            rows = []
            for item in items:
                row = []
                for c in cols:
                    val = item.get(c, "")
                    if c == "integrity_flags" and val:
                        val = "FLAG: " + ", ".join(val)
                    row.append(val)
                rows.append(row)
            r = write_table(ws_e, r, headers, rows) + 2

    wb.save(out_path)
    print(str(out_path))


if __name__ == "__main__":
    main()
